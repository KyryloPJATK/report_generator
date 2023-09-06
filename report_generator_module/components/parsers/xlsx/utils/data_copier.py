from copy import copy

from openpyxl.cell import Cell, MergedCell
from openpyxl.worksheet.worksheet import Worksheet


class ExcelDataCopier:

    def __init__(self, ws_in: Worksheet, ws_out: Worksheet, wb_out):
        """
            :param ws_in source worksheet
            :param ws_out target worksheet
        """
        self.ws_in = ws_in
        self.ws_out = ws_out
        self.wb_out = wb_out

    def copy_cell(self, source_cell, dest_row: int, dest_col: int = None) -> Cell:
        """
            Copies cell to the destination worksheet

            :param source_cell: Cell object to copy
            :param dest_row: target row index (if None - copied index from source cell)
            :param dest_col: target column index (if None - copied index from source cell)

            :returns Composed cell object
        """
        new_cell = self.ws_out.cell(row=dest_row,
                                    column=dest_col or source_cell.col_idx,
                                    value=source_cell.value)
        if not isinstance(new_cell, MergedCell):
            new_cell.value = source_cell.value
            new_cell.hyperlink = source_cell.hyperlink
            new_cell.number_format = source_cell.number_format
            if source_cell.has_style:
                new_cell._style = copy(source_cell._style)
            return new_cell

    def copy_merged_cells_in_row(self, source_row_idx, dest_row_idx):
        """ Copies all the merged cells that start at the provided source row id into destination worksheet rows"""
        for merged_cell in [cell for cell in self.ws_in.merged_cells if cell.min_row == source_row_idx]:
            row_diff = merged_cell.max_row - merged_cell.min_row
            start_col, end_col = merged_cell.coord.split(':')
            # new_coords = start_col.replace(str(source_row_idx), str(dest_row_idx)) + ':' + \
            #              end_col.replace(str(source_row_idx+row_diff), str(dest_row_idx+row_diff))
            self.ws_out.merge_cells(start_row=dest_row_idx, start_column=merged_cell.min_col,
                                    end_row=dest_row_idx+row_diff, end_column=merged_cell.max_col)

    def copy_row(self, source_row_idx, dest_row_idx, source_row=None):
        """ Full copy of row from one worksheet to another """
        if not source_row:
            source_row = list(self.ws_in.rows)[source_row-1]
        for cell in source_row:
            if not isinstance(cell, MergedCell):
                self.copy_cell(cell, dest_row=dest_row_idx)
        self.copy_merged_cells_in_row(source_row_idx, dest_row_idx)
