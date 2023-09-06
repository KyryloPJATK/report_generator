import copy
import logging
import os
import re
from typing import List, Union, Tuple

import pandas as pd
from openpyxl.cell import Cell, MergedCell
from openpyxl.reader.excel import load_workbook
from openpyxl.workbook import Workbook

from report_generator_module.components.parsers.base.parser import BaseParser
from report_generator_module.components.parsers.base.utils.exceptions import ParserException
from report_generator_module.components.parsers.base.utils.parsing_utils import get_real_value
from report_generator_module.components.parsers.xlsx.strategy.navigator import ExcelParsingStrategyNavigator
from report_generator_module.components.parsers.xlsx.utils.data_copier import ExcelDataCopier
from report_generator_module.components.parsers.xlsx.utils.excel_constants import range_pattern

logger = logging.getLogger(__name__)


class ExcelParser(BaseParser):

    def __init__(self, config):
        super().__init__(config=config)
        self.sheet_name = self.config.get("source_sheet_name", "Sheet1")
        self.wb_in = load_workbook(filename=self.source_path)
        self.ws_in = self.wb_in[self.sheet_name]

    def init_strategy_navigator(self) -> ExcelParsingStrategyNavigator:
        return ExcelParsingStrategyNavigator()

    def init_script_lines(self) -> List[str]:
        return list(pd.read_excel(self.source_path, header=None, sheet_name=self.sheet_name).iloc[:, 0])

    @staticmethod
    def replace_formula_row_idx(formula_str: str, context: dict) -> str:
        """ Replaces formula row based on the context """
        # We ignore recalculation of the formula if it has $sign attached to raw
        row_idx_mapping = context.get('row_idx_mapping', {})
        if '$' not in formula_str[1:]:
            formula_row_idx = int(re.findall(r"[0-9]+", formula_str)[0])
            max_mapped_index = max(list(row_idx_mapping))
            # If selected row is bigger than max item in mapping - adjusting it based on delta
            if formula_row_idx > max_mapped_index:
                max_delta = row_idx_mapping[max_mapped_index] - max_mapped_index
                formula_row_idx += max_delta
            return formula_str[0] + str(context.get('row_idx_mapping', {}).get(int(formula_row_idx), formula_row_idx))
        return formula_str

    @staticmethod
    def reformat_formula_cell(cell, context):
        """
            Reformat cell formula range based on the parser lines index mapping
        """
        # TODO: make insertion-like behavior so that formula cells will be recalculated
        formula = cell.value
        # Find all ranges in the formula using the regular expression
        ranges = re.findall(range_pattern, formula)
        # Print the ranges found in the formula
        if len(ranges) > 0:
            for _range in ranges:
                start_idx, end_idx = _range
                start_idx_new = ExcelParser.replace_formula_row_idx(start_idx, context)
                end_idx_new = ExcelParser.replace_formula_row_idx(end_idx, context)
                cell.value = cell.value.replace(f'{start_idx}:{end_idx}', f'{start_idx_new}:{end_idx_new}')
        else:
            formula_value = formula.split('=')[1]
            cell.value = cell.value.replace(formula_value, ExcelParser.replace_formula_row_idx(formula_value, context))
        print(f"Cell {cell.coordinate} value after range parsing: {cell.value}")
        return cell

    def compose_populated_row(self, context: dict) -> Tuple[Cell]:
        """ Composes excel row with filled values based on the context """
        source_row: Tuple[Union[Cell, MergedCell]] = list(self.ws_in.rows)[context['source_line_idx']-1]
        target_row = []
        # Patch: setting lexeme containing value to None to prevent copying script
        source_row[0].value = None
        for cell in source_row:
            cell_copy = copy.copy(cell)
            if isinstance(cell_copy, Cell):
                cell_copy.value = get_real_value(context, cell_copy.value)
                if cell_copy.data_type == 'f':
                    cell_copy = self.reformat_formula_cell(cell=cell_copy, context=context)
            target_row.append(cell_copy)
        return *target_row,

    def post_lexeme_parse(self, context: dict):
        context['dest_line_idx'] += 1
        context['row_idx_mapping'][context['source_line_idx']] = context['dest_line_idx']
        source_row = self.compose_populated_row(context=context)
        context['data_copier'].copy_row(source_row_idx=context['source_line_idx'],
                                        dest_row_idx=context['dest_line_idx'],
                                        source_row=source_row)

    def parse(self, context: dict = None) -> Union[ParserException, None]:
        """
            Parses logic stored in the source excel-compatible file and outputs it to the destination excel-compatible file

            :param context: parse context
        """
        context = context or {}
        destination_path = context['destination_path']
        destination_sheet_name = context.get('destination_sheet_name', 'Sheet1')

        if not os.path.exists(destination_path):
            wb_out = Workbook()
            wb_out.create_sheet(destination_sheet_name)
            wb_out.save(destination_path)
        context['wb_out'] = load_workbook(destination_path)
        if destination_sheet_name in context['wb_out'].sheetnames:
            ws_out = context['wb_out'][destination_sheet_name]
        else:
            ws_out = context['wb_out'].create_sheet(destination_sheet_name)
        context['data_copier'] = ExcelDataCopier(ws_in=self.ws_in, ws_out=ws_out, wb_out=context['wb_out'])
        return super().parse(context)
